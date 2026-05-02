import sys, os, re
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font, Alignment
import matplotlib
matplotlib.use("QtAgg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QLineEdit, QComboBox, QTextEdit, QTabWidget,
    QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox, QDialog,
    QGridLayout, QScrollArea, QFrame, QCalendarWidget
)
from PyQt6.QtCore import Qt, pyqtSignal, QDate
from PyQt6.QtGui import QFont, QColor, QPalette, QTextCharFormat

# ── Excel file path (same folder as exe) ──────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(sys.argv[0]))
EXCEL_FILE = os.path.join(BASE_DIR, "Trade_Diary_Backtest_1.xlsx")

COLUMNS = [
    "Date", "Day", "Entry ", "Index", "Entry with Rules",
    "Long or short", "Result", "Second Trade ",
    "Points Captured / Loss", "Loss / No trade  reason"
]

C = {
    "bg": "#FAF8F5", "bg2": "#F2EFE9", "bg3": "#E8E3DB",
    "white": "#FFFFFF", "ink": "#1A1A1A", "ink2": "#444444", "ink3": "#888888",
    "accent": "#C8512A", "accent_l": "#F5ECE7",
    "green": "#2A7A4E", "green_l": "#E7F4ED",
    "red": "#B83232", "red_l": "#FAEAEA",
    "gold": "#B8882A", "gold_l": "#FDF5E7",
    "border": "#D9D3C9",
}

# ── Excel helpers ─────────────────────────────────────────────────────────────
HEADER_FILL  = "BDD7EE"   # original blue header colour
TABLE_STYLE  = "TableStyleLight13"
TABLE_NAME   = "Table1"
TABLE_ROWS   = 10000       # pre-allocated table range (matches original file)

def load_df():
    if os.path.exists(EXCEL_FILE):
        df = pd.read_excel(EXCEL_FILE, dtype=str)
        df.fillna("", inplace=True)
        for col in COLUMNS:
            if col not in df.columns:
                df[col] = ""
        return df[COLUMNS].copy()
    return pd.DataFrame(columns=COLUMNS)

def save_df(df):
    """Write df back to Excel preserving the original Table format."""
    df = df[COLUMNS].copy()
    data_rows = len(df)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ── Write header ──
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(name="Calibri", size=11)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Write data rows ──
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            val = row[col_name]
            ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)

    # ── Column widths (approximate originals) ──
    col_widths = [12, 12, 8, 10, 16, 14, 10, 14, 22, 28]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Recreate the Excel Table over A1:J{TABLE_ROWS} ──
    last_col_letter = openpyxl.utils.get_column_letter(len(COLUMNS))
    table_ref = f"A1:{last_col_letter}{TABLE_ROWS}"
    tbl = Table(displayName=TABLE_NAME, ref=table_ref)
    style = TableStyleInfo(
        name=TABLE_STYLE,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tbl.tableStyleInfo = style
    ws.add_table(tbl)

    wb.save(EXCEL_FILE)

def init_excel():
    """Create a fresh Excel file with the proper Table format if none exists."""
    save_df(pd.DataFrame(columns=COLUMNS))

def parse_date_str(s):
    try:
        p = str(s).strip().split("/")
        return datetime(int(p[2]), int(p[1]), int(p[0]))
    except:
        return datetime.min

def sort_df(df):
    df = df.copy()
    df["_sort"] = df["Date"].apply(parse_date_str)
    return df.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)

def day_from_date(s):
    try:
        p = str(s).strip().split("/")
        return datetime(int(p[2]), int(p[1]), int(p[0])).strftime("%A")
    except:
        return ""

def is_valid_date(s):
    if not s or len(s) != 10:
        return False
    try:
        p = s.split("/")
        if len(p) != 3 or len(p[0]) != 2 or len(p[1]) != 2 or len(p[2]) != 4:
            return False
        datetime(int(p[2]), int(p[1]), int(p[0]))
        return True
    except:
        return False

def is_valid_points(v):
    if not v or v == "-":
        return True
    return bool(re.match(r'^-?\d+(\.\d+)?$', v.strip()))

# ── UI helpers ────────────────────────────────────────────────────────────────
def cap_label(text):
    lbl = QLabel(text.upper())
    lbl.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
    lbl.setStyleSheet(f"color:{C['ink3']};letter-spacing:1px;background:transparent;border:none;")
    return lbl

def divider():
    f = QFrame()
    f.setFrameShape(QFrame.Shape.HLine)
    f.setStyleSheet(f"background:{C['border']};max-height:1px;border:none;")
    return f

def make_combo(options):
    cb = QComboBox()
    cb.addItem("— Select —", "")
    for o in options:
        cb.addItem(o, o)
    cb.setStyleSheet(f"""
        QComboBox{{border:1px solid {C['border']};border-radius:6px;padding:8px 12px;
            background:{C['bg']};color:{C['ink']};font-family:Segoe UI;font-size:13px;min-height:36px;}}
        QComboBox:focus{{border-color:{C['accent']};background:white;}}
        QComboBox::drop-down{{border:none;width:24px;}}
        QComboBox QAbstractItemView{{border:1px solid {C['border']};background:white;
            selection-background-color:{C['accent_l']};selection-color:{C['ink']};
            font-family:Segoe UI;font-size:13px;}}
    """)
    return cb

def make_lineedit(placeholder="", readonly=False):
    le = QLineEdit()
    le.setPlaceholderText(placeholder)
    le.setFixedHeight(38)
    le.setReadOnly(readonly)
    bg = C['bg2'] if readonly else C['bg']
    col = C['ink2'] if readonly else C['ink']
    le.setStyleSheet(f"""
        QLineEdit{{border:1px solid {C['border']};border-radius:6px;padding:8px 12px;
            background:{bg};color:{col};font-family:Segoe UI;font-size:13px;}}
        QLineEdit:focus{{border-color:{C['accent']};background:white;}}
        QLineEdit:read-only{{background:{C['bg2']};color:{C['ink2']};}}
    """)
    return le

def make_btn(text, color="accent", small=False):
    btn = QPushButton(text)
    h = 32 if small else 40
    pad = "5px 12px" if small else "8px 20px"
    fs = 11 if small else 13
    bgs = {"accent":C["accent"],"green":C["green"],"red":C["red"],"ghost":C["white"]}
    hvs = {"accent":"#B0431E","green":"#1D6040","red":"#9A2828","ghost":C["bg2"]}
    bg, hv = bgs.get(color,C["accent"]), hvs.get(color,C["accent"])
    fg = C["ink"] if color=="ghost" else "white"
    bdr = f"border:1px solid {C['border']};" if color=="ghost" else "border:none;"
    btn.setStyleSheet(f"""
        QPushButton{{background:{bg};color:{fg};{bdr}border-radius:6px;
            padding:{pad};font-family:Segoe UI;font-size:{fs}px;font-weight:600;min-height:{h}px;}}
        QPushButton:hover{{background:{hv};}}
        QPushButton:disabled{{background:{C['bg3']};color:{C['ink3']};}}
    """)
    return btn

def make_textarea():
    te = QTextEdit()
    te.setFixedHeight(76)
    te.setStyleSheet(f"""
        QTextEdit{{border:1px solid {C['border']};border-radius:6px;padding:6px;
            background:{C['bg']};color:{C['ink']};font-family:Segoe UI;font-size:13px;}}
        QTextEdit:focus{{border-color:{C['accent']};background:white;}}
    """)
    return te

def set_combo_value(cb, val):
    idx = cb.findText(str(val))
    cb.setCurrentIndex(idx if idx >= 0 else 0)

def warn_box(parent, title, msg):
    QMessageBox.warning(parent, title, msg)

def info_box(parent, title, msg):
    QMessageBox.information(parent, title, msg)

# ── Calendar Picker Dialog ─────────────────────────────────────────────────────
class CalendarDialog(QDialog):
    def __init__(self, parent, current_date_str=""):
        super().__init__(parent)
        self.setWindowTitle("Pick a Date")
        self.setFixedSize(340, 310)
        self.setStyleSheet(f"""
            QDialog {{ background:{C['white']}; }}
            QCalendarWidget QAbstractItemView {{
                font-family: Segoe UI; font-size: 13px;
                selection-background-color: {C['accent']};
                selection-color: white;
                background: white;
                color: {C['ink']};
            }}
            QCalendarWidget QWidget#qt_calendar_navigationbar {{
                background: {C['ink']}; border-radius: 8px 8px 0 0;
                padding: 4px;
            }}
            QCalendarWidget QToolButton {{
                color: white; background: transparent; border: none;
                font-family: Segoe UI; font-size: 13px; font-weight: bold;
                padding: 4px 8px;
            }}
            QCalendarWidget QToolButton:hover {{
                background: rgba(255,255,255,0.15); border-radius: 4px;
            }}
            QCalendarWidget QMenu {{
                background: white; border: 1px solid {C['border']};
                font-family: Segoe UI; font-size: 12px; color: {C['ink']};
            }}
            QCalendarWidget QSpinBox {{
                color: white; background: transparent; border: none;
                font-family: Segoe UI; font-size: 13px; font-weight: bold;
            }}
            QCalendarWidget QSpinBox::up-button, QCalendarWidget QSpinBox::down-button {{
                width: 0px;
            }}
            QCalendarWidget QAbstractItemView:enabled {{
                color: {C['ink']}; background: white;
            }}
            QCalendarWidget QAbstractItemView:disabled {{
                color: {C['ink3']};
            }}
        """)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(12, 12, 12, 12)
        lay.setSpacing(10)

        self.cal = QCalendarWidget()
        self.cal.setGridVisible(False)
        self.cal.setVerticalHeaderFormat(QCalendarWidget.VerticalHeaderFormat.NoVerticalHeader)
        self.cal.setHorizontalHeaderFormat(QCalendarWidget.HorizontalHeaderFormat.ShortDayNames)

        # Style weekend header text (Sat/Sun) same as weekdays
        fmt = QTextCharFormat()
        fmt.setForeground(QColor(C['ink']))
        self.cal.setWeekdayTextFormat(Qt.DayOfWeek.Saturday, fmt)
        self.cal.setWeekdayTextFormat(Qt.DayOfWeek.Sunday, fmt)

        # Pre-select date if valid
        if current_date_str and is_valid_date(current_date_str):
            p = current_date_str.split("/")
            self.cal.setSelectedDate(QDate(int(p[2]), int(p[1]), int(p[0])))

        self.cal.activated.connect(self.accept)   # double-click selects
        lay.addWidget(self.cal)

        btn_row = QHBoxLayout(); btn_row.setSpacing(8)
        cancel = make_btn("Cancel", "ghost", small=True)
        cancel.clicked.connect(self.reject)
        select = make_btn("Select Date", "accent", small=True)
        select.clicked.connect(self.accept)
        btn_row.addStretch(); btn_row.addWidget(cancel); btn_row.addWidget(select)
        lay.addLayout(btn_row)

    def selected_date_str(self):
        d = self.cal.selectedDate()
        return f"{d.day():02d}/{d.month():02d}/{d.year()}"

# ── Date row: input + calendar button together ────────────────────────────────
def make_date_row(date_edit, parent_widget):
    """Returns an HBoxLayout containing the date lineedit + a calendar icon button."""
    row = QHBoxLayout(); row.setSpacing(6); row.setContentsMargins(0,0,0,0)
    row.addWidget(date_edit)
    cal_btn = QPushButton("📅")
    cal_btn.setFixedSize(38, 38)
    cal_btn.setToolTip("Pick date from calendar")
    cal_btn.setStyleSheet(f"""
        QPushButton {{
            background: {C['accent_l']}; border: 1px solid {C['border']};
            border-radius: 6px; font-size: 16px; padding: 0;
        }}
        QPushButton:hover {{ background: {C['accent']}; }}
    """)
    def open_cal():
        dlg = CalendarDialog(parent_widget, date_edit.text().strip())
        if dlg.exec():
            date_edit.setText(dlg.selected_date_str())
    cal_btn.clicked.connect(open_cal)
    row.addWidget(cal_btn)
    return row

# ── Confirm Dialog ─────────────────────────────────────────────────────────────
class ConfirmDialog(QDialog):
    def __init__(self, parent, title, message, btn_label="Confirm", danger=False):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setFixedWidth(400)
        self.setStyleSheet("background:white;font-family:Segoe UI;")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(24,24,24,18)
        lay.setSpacing(12)
        t = QLabel(title)
        t.setFont(QFont("Segoe UI",14,QFont.Weight.Bold))
        t.setStyleSheet("background:transparent;border:none;")
        lay.addWidget(t)
        w = QLabel(message)
        w.setWordWrap(True)
        w.setStyleSheet("background:#FFF8E7;border:1px solid #F0D070;border-radius:6px;padding:12px;color:#7A5500;font-size:13px;")
        lay.addWidget(w)
        row = QHBoxLayout()
        row.setSpacing(8)
        c = make_btn("Cancel","ghost",small=True)
        c.clicked.connect(self.reject)
        ok = make_btn(btn_label,"red" if danger else "green",small=True)
        ok.clicked.connect(self.accept)
        row.addStretch(); row.addWidget(c); row.addWidget(ok)
        lay.addLayout(row)

# ── Edit Dialog ────────────────────────────────────────────────────────────────
class EditDialog(QDialog):
    def __init__(self, parent, date, row):
        super().__init__(parent)
        self.date = date
        self.setWindowTitle(f"Edit — {date}")
        self.setFixedWidth(540)
        self.setStyleSheet("background:white;font-family:Segoe UI;")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(24,22,24,18)
        lay.setSpacing(12)

        t = QLabel(f"Edit Trade — {date}")
        t.setFont(QFont("Segoe UI",14,QFont.Weight.Bold))
        t.setStyleSheet("background:transparent;border:none;")
        lay.addWidget(t)
        lay.addWidget(divider())

        g = QGridLayout(); g.setSpacing(10)
        g.setColumnStretch(0,1); g.setColumnStretch(1,1)

        def add_cb(label, r, c, opts, cur):
            g.addWidget(cap_label(label), r*2, c)
            cb = make_combo(opts); set_combo_value(cb, cur)
            g.addWidget(cb, r*2+1, c); return cb

        self.entry_cb     = add_cb("Entry",          0,0,["Yes","No"],                        row["Entry "])
        self.index_cb     = add_cb("Index",          0,1,["Nifty","Sensex","Forex","Gold","Other"],row["Index"])
        self.rules_cb     = add_cb("Entry w/ Rules", 1,0,["Yes","No","No Trade"],             row["Entry with Rules"])
        self.direction_cb = add_cb("Long or Short",  1,1,["Long","Short","No Trade"],         row["Long or short"])
        self.result_cb    = add_cb("Result",         2,0,["Profit","Loss","No Trade","CTC"],  row["Result"])
        self.second_cb    = add_cb("Second Trade",   2,1,["Yes","No"],                        row["Second Trade "])

        g.addWidget(cap_label("Points"), 6,0)
        self.points_edit = make_lineedit()
        self.points_edit.setText(str(row["Points Captured / Loss"]))
        g.addWidget(self.points_edit, 7,0)

        g.addWidget(cap_label("Reason"), 6,1)
        self.reason_edit = make_textarea()
        self.reason_edit.setPlainText(str(row["Loss / No trade  reason"]))
        g.addWidget(self.reason_edit, 7,1)
        lay.addLayout(g)
        lay.addWidget(divider())

        row2 = QHBoxLayout(); row2.setSpacing(8)
        c = make_btn("Cancel","ghost",small=True); c.clicked.connect(self.reject)
        ok = make_btn("Confirm Changes","green",small=True); ok.clicked.connect(self._confirm)
        row2.addStretch(); row2.addWidget(c); row2.addWidget(ok)
        lay.addLayout(row2)

    def _confirm(self):
        pts = self.points_edit.text().strip()
        if not is_valid_points(pts):
            warn_box(self,"Invalid","Points must be a number."); return
        dlg = ConfirmDialog(self,"Save Changes?",
            "Save these changes?\nThe Excel file will be updated immediately.","Save to Excel")
        if dlg.exec(): self.accept()

    def get_data(self):
        return {
            "Date":                    self.date,
            "Day":                     day_from_date(self.date),
            "Entry ":                  self.entry_cb.currentData() or "",
            "Index":                   self.index_cb.currentData() or "",
            "Entry with Rules":        self.rules_cb.currentData() or "",
            "Long or short":           self.direction_cb.currentData() or "",
            "Result":                  self.result_cb.currentData() or "",
            "Second Trade ":           self.second_cb.currentData() or "",
            "Points Captured / Loss":  self.points_edit.text().strip(),
            "Loss / No trade  reason": self.reason_edit.toPlainText().strip(),
        }

# ── Shared: open edit/delete helpers ─────────────────────────────────────────
def do_edit(parent, date, callback):
    df = load_df()
    match = df[df["Date"] == date]
    if match.empty: return
    dlg = EditDialog(parent, date, match.iloc[0])
    if dlg.exec():
        data = dlg.get_data()
        dff = load_df()
        dff = dff[dff["Date"] != date]
        dff = sort_df(pd.concat([dff, pd.DataFrame([data])], ignore_index=True))
        save_df(dff)
        callback()
        info_box(parent,"Updated",f"Trade for {date} updated in Excel.")

def do_delete(parent, date, callback):
    dlg = ConfirmDialog(parent,"Remove Trade",
        f"Remove trade for {date}?\nThis cannot be undone.","Remove",danger=True)
    if dlg.exec():
        df = load_df()
        df = df[df["Date"] != date]
        save_df(df)
        callback()
        info_box(parent,"Removed",f"Trade for {date} removed from Excel.")

# ── Tab: Add Trade ─────────────────────────────────────────────────────────────
class AddTradeTab(QWidget):
    trade_saved = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.setStyleSheet(f"background:{C['bg']};")
        outer = QVBoxLayout(self)
        outer.setContentsMargins(28,24,28,24); outer.setSpacing(0)

        card = QWidget()
        card.setStyleSheet(f"background:{C['white']};border:1px solid {C['border']};border-radius:10px;")
        cl = QVBoxLayout(card); cl.setContentsMargins(28,24,28,28); cl.setSpacing(14)

        t = QLabel("Log a New Trade")
        t.setFont(QFont("Segoe UI",15,QFont.Weight.Bold))
        t.setStyleSheet(f"color:{C['ink']};border-bottom:1px solid {C['bg3']};padding-bottom:12px;background:transparent;border-left:none;border-right:none;border-top:none;")
        cl.addWidget(t)

        g = QGridLayout(); g.setSpacing(14)
        g.setColumnStretch(0,1); g.setColumnStretch(1,1)

        g.addWidget(cap_label("Date"),0,0)
        self.date_edit = make_lineedit("DD/MM/YYYY")
        self.date_edit.setMaxLength(10)
        self.date_edit.textChanged.connect(self._on_date)
        g.addLayout(make_date_row(self.date_edit, self),1,0)

        g.addWidget(cap_label("Day (auto)"),0,1)
        self.day_edit = make_lineedit("Auto-filled",readonly=True)
        g.addWidget(self.day_edit,1,1)

        g.addWidget(cap_label("Entry"),2,0)
        self.entry_cb = make_combo(["Yes","No"])
        g.addWidget(self.entry_cb,3,0)

        g.addWidget(cap_label("Index"),2,1)
        self.index_cb = make_combo(["Nifty","Sensex","Forex","Gold","Other"])
        g.addWidget(self.index_cb,3,1)

        g.addWidget(cap_label("Entry with Rules"),4,0)
        self.rules_cb = make_combo(["Yes","No","No Trade"])
        g.addWidget(self.rules_cb,5,0)

        g.addWidget(cap_label("Long or Short"),4,1)
        self.direction_cb = make_combo(["Long","Short","No Trade"])
        g.addWidget(self.direction_cb,5,1)

        g.addWidget(cap_label("Result"),6,0)
        self.result_cb = make_combo(["Profit","Loss","No Trade","CTC"])
        g.addWidget(self.result_cb,7,0)

        g.addWidget(cap_label("Second Trade"),6,1)
        self.second_cb = make_combo(["Yes","No"])
        g.addWidget(self.second_cb,7,1)

        g.addWidget(cap_label("Points Captured / Loss"),8,0)
        self.points_edit = make_lineedit("e.g. 77 or -23")
        self.points_edit.textChanged.connect(self._validate_points)
        g.addWidget(self.points_edit,9,0)

        self.points_err = QLabel("⚠  Only numbers are allowed")
        self.points_err.setStyleSheet(f"color:{C['red']};font-size:11px;background:transparent;border:none;")
        self.points_err.hide()
        g.addWidget(self.points_err,10,0)

        g.addWidget(cap_label("Loss / No Trade Reason"),8,1)
        self.reason_edit = make_textarea()
        g.addWidget(self.reason_edit,9,1,2,1)

        cl.addLayout(g)
        cl.addWidget(divider())

        brow = QHBoxLayout(); brow.setSpacing(10)
        self.submit_btn = make_btn("✓  Submit Trade","accent")
        self.submit_btn.clicked.connect(self.submit_trade)
        clr = make_btn("Clear","ghost"); clr.clicked.connect(self.clear_form)
        brow.addWidget(self.submit_btn); brow.addWidget(clr); brow.addStretch()
        cl.addLayout(brow)

        outer.addWidget(card); outer.addStretch()

    def _on_date(self, text):
        digits = text.replace("/","")
        fmt = ""
        for i,ch in enumerate(digits[:8]):
            if i in (2,4): fmt += "/"
            fmt += ch
        if fmt != text:
            self.date_edit.blockSignals(True)
            self.date_edit.setText(fmt)
            self.date_edit.blockSignals(False)
        self.day_edit.setText(day_from_date(fmt))

    def _validate_points(self):
        v = self.points_edit.text().strip()
        ok = is_valid_points(v)
        self.points_err.setVisible(not ok)
        return ok

    def clear_form(self):
        for w in [self.date_edit,self.day_edit,self.points_edit]:
            w.clear()
        self.reason_edit.clear()
        for cb in [self.entry_cb,self.index_cb,self.rules_cb,
                   self.direction_cb,self.result_cb,self.second_cb]:
            cb.setCurrentIndex(0)
        self.points_err.hide()

    def submit_trade(self):
        date = self.date_edit.text().strip()
        if not is_valid_date(date):
            warn_box(self,"Invalid Date","Please enter a valid date in DD/MM/YYYY format."); return
        if not self._validate_points():
            warn_box(self,"Invalid Points","Points field must contain numbers only."); return

        df = load_df()
        if not df[df["Date"]==date].empty:
            r = QMessageBox.question(self,"Overwrite?",
                f"A trade for {date} already exists.\nOverwrite it?",
                QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No)
            if r != QMessageBox.StandardButton.Yes: return

        df = df[df["Date"]!=date]
        new = pd.DataFrame([{
            "Date":date,"Day":day_from_date(date),
            "Entry ":self.entry_cb.currentData() or "",
            "Index":self.index_cb.currentData() or "",
            "Entry with Rules":self.rules_cb.currentData() or "",
            "Long or short":self.direction_cb.currentData() or "",
            "Result":self.result_cb.currentData() or "",
            "Second Trade ":self.second_cb.currentData() or "",
            "Points Captured / Loss":self.points_edit.text().strip(),
            "Loss / No trade  reason":self.reason_edit.toPlainText().strip(),
        }])
        save_df(sort_df(pd.concat([df,new],ignore_index=True)))
        self.clear_form()
        self.trade_saved.emit()
        info_box(self,"Saved",f"Trade for {date} saved to Excel!")

# ── Tab: Find by Date ─────────────────────────────────────────────────────────
class FindTab(QWidget):
    refresh_needed = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.setStyleSheet(f"background:{C['bg']};")
        outer = QVBoxLayout(self)
        outer.setContentsMargins(28,24,28,24); outer.setSpacing(14)

        sc = QWidget()
        sc.setStyleSheet(f"background:{C['white']};border:1px solid {C['border']};border-radius:10px;")
        sl = QVBoxLayout(sc); sl.setContentsMargins(24,18,24,18); sl.setSpacing(10)
        t = QLabel("Find Trade by Date")
        t.setFont(QFont("Segoe UI",15,QFont.Weight.Bold))
        t.setStyleSheet("background:transparent;border:none;")
        sl.addWidget(t)
        row = QHBoxLayout(); row.setSpacing(10)
        self.search_edit = make_lineedit("DD/MM/YYYY")
        self.search_edit.setMaxLength(10)
        self.search_edit.textChanged.connect(self._fmt)
        self.search_edit.returnPressed.connect(self.find_trade)
        date_row = make_date_row(self.search_edit, self)
        row.addLayout(date_row)
        sb = make_btn("Search","accent"); sb.clicked.connect(self.find_trade)
        row.addWidget(sb)
        sl.addLayout(row)
        outer.addWidget(sc)

        self.result_card = QWidget()
        self.result_card.setStyleSheet(f"background:{C['white']};border:1px solid {C['border']};border-radius:10px;")
        self.rl = QVBoxLayout(self.result_card)
        self.rl.setContentsMargins(24,18,24,18); self.rl.setSpacing(10)
        self.result_card.hide()
        outer.addWidget(self.result_card)
        outer.addStretch()

    def _fmt(self, text):
        digits = text.replace("/","")
        fmt = ""
        for i,ch in enumerate(digits[:8]):
            if i in (2,4): fmt += "/"
            fmt += ch
        if fmt != text:
            self.search_edit.blockSignals(True)
            self.search_edit.setText(fmt)
            self.search_edit.blockSignals(False)

    def _clear_result(self):
        while self.rl.count():
            item = self.rl.takeAt(0)
            if item.widget(): item.widget().deleteLater()

    def find_trade(self):
        date = self.search_edit.text().strip()
        if not is_valid_date(date):
            warn_box(self,"Invalid Date","Please enter a valid date."); return
        df = load_df()
        match = df[df["Date"]==date]
        self._clear_result()
        if match.empty:
            lbl = QLabel(f"No trade found for  {date}")
            lbl.setStyleSheet(f"color:{C['ink3']};font-size:13px;background:transparent;border:none;")
            self.rl.addWidget(lbl)
            self.result_card.show(); return

        r = match.iloc[0]
        hrow = QHBoxLayout()
        dl = QLabel(date)
        dl.setFont(QFont("Segoe UI Semibold",14))
        dl.setStyleSheet("background:transparent;border:none;")
        hrow.addWidget(dl); hrow.addStretch()
        eb = make_btn("✏  Edit","ghost",small=True)
        eb.clicked.connect(lambda: do_edit(self, date, self._after_edit))
        rb = make_btn("🗑  Remove","red",small=True)
        rb.clicked.connect(lambda: do_delete(self, date, self._after_delete))
        hrow.addWidget(eb); hrow.addWidget(rb)
        self.rl.addLayout(hrow)
        self.rl.addWidget(divider())

        fields = [
            ("Day",r["Day"]),("Entry",r["Entry "]),("Index",r["Index"]),
            ("Rules",r["Entry with Rules"]),("Direction",r["Long or short"]),
            ("Result",r["Result"]),("2nd Trade",r["Second Trade "]),
            ("Points",r["Points Captured / Loss"] or "—"),
            ("Reason",r["Loss / No trade  reason"] or "—"),
        ]
        row_w = QHBoxLayout(); row_w.setSpacing(8); count = 0
        for k,v in fields:
            pill = QWidget()
            pill.setStyleSheet(f"background:{C['bg']};border:1px solid {C['border']};border-radius:6px;")
            pl = QVBoxLayout(pill); pl.setContentsMargins(10,6,10,6); pl.setSpacing(2)
            kl = QLabel(k.upper())
            kl.setStyleSheet(f"color:{C['ink3']};font-size:9px;font-weight:bold;background:transparent;border:none;")
            vl = QLabel(str(v) if v else "—")
            vl.setStyleSheet(f"color:{C['ink']};font-size:12px;font-weight:600;background:transparent;border:none;")
            pl.addWidget(kl); pl.addWidget(vl)
            row_w.addWidget(pill); count += 1
            if count == 5:
                self.rl.addLayout(row_w)
                row_w = QHBoxLayout(); row_w.setSpacing(8)
        if count % 5: self.rl.addLayout(row_w)
        self.result_card.show()

    def _after_edit(self):
        self.refresh_needed.emit()
        self.find_trade()

    def _after_delete(self):
        self.refresh_needed.emit()
        self._clear_result()
        self.result_card.hide()
        self.search_edit.clear()

# ── Tab: Check Results ─────────────────────────────────────────────────────────
class ResultsTab(QWidget):
    refresh_needed = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.current_filter = "all"
        self.setStyleSheet(f"background:{C['bg']};")
        outer = QVBoxLayout(self)
        outer.setContentsMargins(28,24,28,24); outer.setSpacing(0)

        card = QWidget()
        card.setStyleSheet(f"background:{C['white']};border:1px solid {C['border']};border-radius:10px;")
        cl = QVBoxLayout(card); cl.setContentsMargins(24,20,24,16); cl.setSpacing(12)

        t = QLabel("Check Results by Category")
        t.setFont(QFont("Segoe UI",15,QFont.Weight.Bold))
        t.setStyleSheet("background:transparent;border:none;")
        cl.addWidget(t)

        frow = QHBoxLayout(); frow.setSpacing(8)
        self.filter_btns = {}
        for label,key in [("All Trades","all"),("Profit","Profit"),("Loss","Loss"),("CTC","CTC"),("No Trade","No Trade")]:
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setStyleSheet(f"""
                QPushButton{{border:1.5px solid {C['border']};border-radius:14px;padding:6px 16px;
                    background:white;color:{C['ink2']};font-family:Segoe UI;font-size:12px;font-weight:600;}}
                QPushButton:checked{{background:{C['accent']};color:white;border-color:{C['accent']};}}
                QPushButton:hover{{border-color:{C['accent']};}}
            """)
            btn.clicked.connect(lambda _,k=key: self.set_filter(k))
            self.filter_btns[key] = btn; frow.addWidget(btn)
        frow.addStretch(); cl.addLayout(frow)
        cl.addWidget(divider())

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea{border:none;background:transparent;}")
        self.rc = QWidget(); self.rc.setStyleSheet("background:transparent;")
        self.rl = QVBoxLayout(self.rc); self.rl.setSpacing(10); self.rl.setContentsMargins(0,0,0,0)
        scroll.setWidget(self.rc)
        cl.addWidget(scroll)
        outer.addWidget(card)
        self.set_filter("all")

    def set_filter(self, key):
        self.current_filter = key
        for k,btn in self.filter_btns.items(): btn.setChecked(k==key)
        self.load_results()

    def load_results(self):
        while self.rl.count():
            item = self.rl.takeAt(0)
            if item.widget(): item.widget().deleteLater()

        df = load_df()
        if self.current_filter != "all":
            df = df[df["Result"]==self.current_filter]

        if df.empty:
            lbl = QLabel("No trades found for this filter.")
            lbl.setStyleSheet(f"color:{C['ink3']};font-size:13px;padding:20px;background:transparent;border:none;")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.rl.addWidget(lbl); self.rl.addStretch(); return

        rc = {"Profit":C["green"],"Loss":C["red"],"CTC":C["gold"],"No Trade":C["ink3"]}
        for _,r in df.iterrows():
            rw = QWidget()
            rw.setStyleSheet(f"background:{C['bg']};border:1px solid {C['border']};border-radius:8px;")
            rwl = QVBoxLayout(rw); rwl.setContentsMargins(16,12,16,12); rwl.setSpacing(8)

            top = QHBoxLayout()
            dl = QLabel(f"{r['Date']}  —  {r['Day']}")
            dl.setFont(QFont("Segoe UI Semibold",13))
            dl.setStyleSheet("background:transparent;border:none;")
            top.addWidget(dl); top.addStretch()
            res = str(r.get("Result",""))
            rc_ = rc.get(res,C["ink3"])
            badge = QLabel(res or "—")
            badge.setStyleSheet(f"background:transparent;border:1.5px solid {rc_};border-radius:10px;padding:2px 10px;color:{rc_};font-size:11px;font-weight:bold;")
            top.addWidget(badge)
            eb = make_btn("✏","ghost",small=True); eb.setFixedWidth(36)
            eb.clicked.connect(lambda _,d=str(r["Date"]): do_edit(self,d,self.load_results))
            rb2 = make_btn("🗑","red",small=True); rb2.setFixedWidth(36)
            rb2.clicked.connect(lambda _,d=str(r["Date"]): do_delete(self,d,self._after))
            top.addWidget(eb); top.addWidget(rb2)
            rwl.addLayout(top)

            pills = QHBoxLayout(); pills.setSpacing(6)
            for k,v in [("Entry",r["Entry "]),("Index",r["Index"]),
                        ("Direction",r["Long or short"]),("Points",r["Points Captured / Loss"])]:
                p = QWidget()
                p.setStyleSheet(f"background:white;border:1px solid {C['border']};border-radius:5px;")
                pl = QVBoxLayout(p); pl.setContentsMargins(8,4,8,4); pl.setSpacing(1)
                kl = QLabel(k.upper())
                kl.setStyleSheet(f"color:{C['ink3']};font-size:9px;font-weight:bold;background:transparent;border:none;")
                vl = QLabel(str(v) if v else "—")
                vl.setStyleSheet(f"color:{C['ink']};font-size:12px;font-weight:600;background:transparent;border:none;")
                pl.addWidget(kl); pl.addWidget(vl); pills.addWidget(p)
            pills.addStretch(); rwl.addLayout(pills)
            self.rl.addWidget(rw)
        self.rl.addStretch()

    def _after(self):
        self.refresh_needed.emit(); self.load_results()

# ── Tab: P/L Ratio ─────────────────────────────────────────────────────────────
class RatioTab(QWidget):
    def __init__(self):
        super().__init__()
        self.setStyleSheet(f"background:{C['bg']};")
        outer = QVBoxLayout(self)
        outer.setContentsMargins(28,24,28,24); outer.setSpacing(0)

        self.card = QWidget()
        self.card.setStyleSheet(f"background:{C['white']};border:1px solid {C['border']};border-radius:10px;")
        self.cl = QVBoxLayout(self.card); self.cl.setContentsMargins(28,24,28,28); self.cl.setSpacing(14)

        t = QLabel("Profit / Loss / CTC Ratio")
        t.setFont(QFont("Segoe UI",15,QFont.Weight.Bold))
        t.setStyleSheet("background:transparent;border:none;")
        self.cl.addWidget(t); self.cl.addWidget(divider())

        self.stat_row = QHBoxLayout(); self.stat_row.setSpacing(16)
        self.cl.addLayout(self.stat_row)
        self.ratio_lbl = QLabel("")
        self.ratio_lbl.setFont(QFont("Segoe UI",13))
        self.ratio_lbl.setStyleSheet(f"color:{C['ink2']};background:transparent;border:none;")
        self.cl.addWidget(self.ratio_lbl)
        self.cl.addStretch()
        outer.addWidget(self.card); outer.addStretch()
        self.load()

    def _stat(self, n, label, bg, fg, bdr):
        w = QWidget()
        w.setStyleSheet(f"background:{bg};border:1px solid {bdr};border-radius:10px;")
        wl = QVBoxLayout(w); wl.setContentsMargins(20,16,20,16)
        nl = QLabel(str(n)); nl.setFont(QFont("Segoe UI",36,QFont.Weight.Bold))
        nl.setStyleSheet(f"color:{fg};background:transparent;border:none;"); nl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ll = QLabel(label.upper())
        ll.setStyleSheet(f"color:{C['ink3']};font-size:10px;font-weight:bold;letter-spacing:1px;background:transparent;border:none;")
        ll.setAlignment(Qt.AlignmentFlag.AlignCenter)
        wl.addWidget(nl); wl.addWidget(ll); return w

    def load(self):
        while self.stat_row.count():
            item = self.stat_row.takeAt(0)
            if item.widget(): item.widget().deleteLater()
        df = load_df()
        p = len(df[df["Result"]=="Profit"])
        l = len(df[df["Result"]=="Loss"])
        c = len(df[df["Result"]=="CTC"])
        tot = p+l+c
        self.stat_row.addWidget(self._stat(p,"Profits",C["green_l"],C["green"],"#B5DDC6"))
        self.stat_row.addWidget(self._stat(l,"Losses", C["red_l"],  C["red"],  "#E5C0C0"))
        self.stat_row.addWidget(self._stat(c,"CTC",    C["gold_l"], C["gold"], "#E5D0A0"))
        if tot:
            self.ratio_lbl.setText(
                f"Ratio — Profit : Loss : CTC = {p} : {l} : {c}"
                f"   ({p/tot*100:.1f}% / {l/tot*100:.1f}% / {c/tot*100:.1f}%)")
        else:
            self.ratio_lbl.setText("No data yet. Add trades to see your ratio.")

# ── Tab: View Sheet ────────────────────────────────────────────────────────────
class SheetTab(QWidget):
    refresh_needed = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.setStyleSheet(f"background:{C['bg']};")
        outer = QVBoxLayout(self)
        outer.setContentsMargins(28,24,28,24); outer.setSpacing(0)

        card = QWidget()
        card.setStyleSheet(f"background:{C['white']};border:1px solid {C['border']};border-radius:10px;")
        cl = QVBoxLayout(card); cl.setContentsMargins(24,20,24,20); cl.setSpacing(12)

        top = QHBoxLayout()
        t = QLabel("All Trade Records")
        t.setFont(QFont("Segoe UI",15,QFont.Weight.Bold))
        t.setStyleSheet("background:transparent;border:none;")
        top.addWidget(t); top.addStretch()
        self.count_lbl = QLabel("")
        self.count_lbl.setStyleSheet(f"color:{C['ink3']};font-size:12px;background:transparent;border:none;")
        top.addWidget(self.count_lbl)
        cl.addLayout(top); cl.addWidget(divider())

        self.table = QTableWidget()
        self.table.setColumnCount(12)
        self.table.setHorizontalHeaderLabels([
            "#","Date","Day","Entry","Index","Rules","Direction","Result","2nd","Points","Reason","Actions"
        ])
        self.table.horizontalHeader().setStyleSheet(f"""
            QHeaderView::section{{background:{C['ink']};color:white;padding:10px 12px;
                font-family:Segoe UI;font-size:11px;font-weight:bold;border:none;letter-spacing:0.5px;}}
        """)
        self.table.setStyleSheet(f"""
            QTableWidget{{border:none;gridline-color:{C['bg3']};background:white;font-family:Segoe UI;font-size:13px;}}
            QTableWidget::item{{padding:8px 12px;border-bottom:1px solid {C['bg3']};color:{C['ink']};}}
            QTableWidget::item:selected{{background:{C['accent_l']};color:{C['ink']};}}
        """)
        self.table.verticalHeader().hide()
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(10, QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(11,110)
        cl.addWidget(self.table)
        outer.addWidget(card)
        self.load()

    def load(self):
        df = load_df()
        self.table.setRowCount(len(df))
        self.count_lbl.setText(f"{len(df)} record{'s' if len(df)!=1 else ''}")
        rc = {"Profit":C["green"],"Loss":C["red"],"CTC":C["gold"],"No Trade":C["ink3"]}
        for i,(_,r) in enumerate(df.iterrows()):
            vals = [str(i+1),r["Date"],r["Day"],r["Entry "],r["Index"],
                    r["Entry with Rules"],r["Long or short"],r["Result"],
                    r["Second Trade "],r["Points Captured / Loss"],r["Loss / No trade  reason"]]
            for j,v in enumerate(vals):
                item = QTableWidgetItem(str(v) if v else "—")
                item.setFont(QFont("Segoe UI",11 if j==0 else 12))
                if j==0: item.setForeground(QColor(C["ink3"]))
                if j==7 and v in rc:
                    item.setForeground(QColor(rc[v]))
                    item.setFont(QFont("Segoe UI",12,QFont.Weight.Bold))
                self.table.setItem(i,j,item)

            aw = QWidget(); al = QHBoxLayout(aw)
            al.setContentsMargins(6,4,6,4); al.setSpacing(6)
            eb = make_btn("✏","ghost",small=True); eb.setFixedWidth(34)
            eb.clicked.connect(lambda _,d=str(r["Date"]): do_edit(self,d,self.load))
            db = make_btn("🗑","red",small=True); db.setFixedWidth(34)
            db.clicked.connect(lambda _,d=str(r["Date"]): do_delete(self,d,self._after))
            al.addWidget(eb); al.addWidget(db)
            self.table.setCellWidget(i,11,aw)
        self.table.resizeRowsToContents()

    def _after(self):
        self.refresh_needed.emit(); self.load()

# ── Tab: Analysis ─────────────────────────────────────────────────────────────
class AnalysisTab(QWidget):

    # ── colour palette used across all charts ─────────────────────────────────
    PROFIT_C  = "#2A7A4E"
    LOSS_C    = "#B83232"
    CTC_C     = "#B8882A"
    YES_C     = "#3A7FC1"
    NO_C      = "#C8512A"
    BG_FIG    = "#FAFAF8"
    BG_AX     = "#FFFFFF"
    TEXT_C    = "#1A1A1A"
    GRID_C    = "#E8E3DB"

    DAY_ORDER = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]

    def __init__(self):
        super().__init__()
        self.setStyleSheet(f"background:{C['bg']};")

        outer = QVBoxLayout(self)
        outer.setContentsMargins(28,24,28,24)
        outer.setSpacing(16)

        # ── Title ──────────────────────────────────────────────────────────────
        title_lbl = QLabel("Analysis Dashboard")
        title_lbl.setFont(QFont("Segoe UI",15,QFont.Weight.Bold))
        title_lbl.setStyleSheet(f"color:{C['ink']};background:transparent;border:none;")
        outer.addWidget(title_lbl)

        # ── Scroll area so charts don't get squashed ───────────────────────────
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea{border:none;background:transparent;}")
        container = QWidget(); container.setStyleSheet("background:transparent;")
        cl = QVBoxLayout(container); cl.setSpacing(20); cl.setContentsMargins(0,0,0,0)
        scroll.setWidget(container)
        outer.addWidget(scroll)

        # ════════════════════════════════════════════════════════════════════
        # SECTION A — PIE CHART
        # ════════════════════════════════════════════════════════════════════
        pie_card = self._make_card()
        pie_card_layout = pie_card.layout()

        pie_header = QHBoxLayout()
        pie_title = QLabel("A — Pie Chart")
        pie_title.setFont(QFont("Segoe UI",13,QFont.Weight.Bold))
        pie_title.setStyleSheet("background:transparent;border:none;")
        pie_header.addWidget(pie_title)
        pie_header.addStretch()

        self.pie_combo = QComboBox()
        for opt in ["Profit / Loss / CTC",
                    "Entry with Rules vs Without Rules",
                    "Trading Days vs Non-Trading Days"]:
            self.pie_combo.addItem(opt)
        self.pie_combo.setStyleSheet(f"""
            QComboBox{{border:1px solid {C['border']};border-radius:6px;padding:6px 12px;
                background:white;color:{C['ink']};font-family:Segoe UI;font-size:13px;min-width:260px;}}
            QComboBox:focus{{border-color:{C['accent']};}}
            QComboBox::drop-down{{border:none;width:22px;}}
            QComboBox QAbstractItemView{{background:white;border:1px solid {C['border']};
                font-family:Segoe UI;font-size:13px;
                selection-background-color:{C['accent_l']};selection-color:{C['ink']};}}
        """)
        self.pie_combo.currentIndexChanged.connect(self._draw_pie)
        pie_header.addWidget(self.pie_combo)
        pie_card_layout.addLayout(pie_header)
        pie_card_layout.addWidget(divider())

        # matplotlib canvas for pie
        self.pie_fig  = Figure(figsize=(9,4.2), facecolor=self.BG_FIG, tight_layout=True)
        self.pie_canvas = FigureCanvas(self.pie_fig)
        self.pie_canvas.setMinimumHeight(340)
        pie_card_layout.addWidget(self.pie_canvas)

        # legend container
        self.pie_legend = QHBoxLayout()
        self.pie_legend.setSpacing(16)
        pie_card_layout.addLayout(self.pie_legend)
        cl.addWidget(pie_card)

        # ════════════════════════════════════════════════════════════════════
        # SECTION B — BAR / STATS CHART
        # ════════════════════════════════════════════════════════════════════
        bar_card = self._make_card()
        bar_card_layout = bar_card.layout()

        bar_header = QHBoxLayout()
        bar_title = QLabel("B — Statistics / Bar Chart")
        bar_title.setFont(QFont("Segoe UI",13,QFont.Weight.Bold))
        bar_title.setStyleSheet("background:transparent;border:none;")
        bar_header.addWidget(bar_title)
        bar_header.addStretch()

        self.bar_combo = QComboBox()
        for opt in ["Points per Trade (by Result)",
                    "Weekly Performance (Day of Week)"]:
            self.bar_combo.addItem(opt)
        self.bar_combo.setStyleSheet(self.pie_combo.styleSheet())
        self.bar_combo.currentIndexChanged.connect(self._draw_bar)
        bar_header.addWidget(self.bar_combo)
        bar_card_layout.addLayout(bar_header)
        bar_card_layout.addWidget(divider())

        self.bar_fig   = Figure(figsize=(9,4.5), facecolor=self.BG_FIG, tight_layout=True)
        self.bar_canvas = FigureCanvas(self.bar_fig)
        self.bar_canvas.setMinimumHeight(360)
        bar_card_layout.addWidget(self.bar_canvas)

        self.bar_legend = QHBoxLayout()
        self.bar_legend.setSpacing(16)
        bar_card_layout.addLayout(self.bar_legend)
        cl.addWidget(bar_card)
        cl.addStretch()

        # ── initial draw ──────────────────────────────────────────────────────
        self.load()

    # ── Card factory ──────────────────────────────────────────────────────────
    def _make_card(self):
        card = QWidget()
        card.setStyleSheet(f"background:{C['white']};border:1px solid {C['border']};border-radius:10px;")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(24,20,24,20)
        lay.setSpacing(14)
        return card

    # ── Legend pill factory ───────────────────────────────────────────────────
    def _clear_legend(self, layout):
        while layout.count():
            item = layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()

    def _add_legend_pill(self, layout, color, label, value=""):
        pill = QWidget()
        pill.setStyleSheet(f"background:{C['bg']};border:1px solid {C['border']};border-radius:6px;")
        pl = QHBoxLayout(pill); pl.setContentsMargins(10,6,12,6); pl.setSpacing(8)
        dot = QLabel("●")
        dot.setStyleSheet(f"color:{color};font-size:14px;background:transparent;border:none;")
        lbl = QLabel(f"{label}  {value}")
        lbl.setStyleSheet(f"color:{C['ink']};font-size:12px;font-weight:600;background:transparent;border:none;")
        pl.addWidget(dot); pl.addWidget(lbl)
        layout.addWidget(pill)

    # ── Shared matplotlib style ───────────────────────────────────────────────
    def _style_ax(self, ax, title=""):
        ax.set_facecolor(self.BG_AX)
        ax.tick_params(colors=self.TEXT_C, labelsize=10)
        ax.spines[["top","right"]].set_visible(False)
        ax.spines[["left","bottom"]].set_color(self.GRID_C)
        ax.yaxis.grid(True, color=self.GRID_C, linewidth=0.8, linestyle="--")
        ax.set_axisbelow(True)
        if title:
            ax.set_title(title, fontsize=12, fontweight="bold",
                         color=self.TEXT_C, pad=10)

    # ══════════════════════════════════════════════════════════════════════════
    # PIE CHART DRAWING
    # ══════════════════════════════════════════════════════════════════════════
    def _draw_pie(self):
        self.pie_fig.clear()
        df = load_df()
        idx = self.pie_combo.currentIndex()
        self._clear_legend(self.pie_legend)

        if df.empty:
            ax = self.pie_fig.add_subplot(111)
            ax.set_facecolor(self.BG_AX)
            ax.text(0.5,0.5,"No data yet — add trades to see charts",
                    ha="center",va="center",fontsize=12,color=self.TEXT_C,
                    transform=ax.transAxes)
            ax.axis("off")
            self.pie_canvas.draw()
            return

        # ── Option 0: Profit / Loss / CTC ─────────────────────────────────
        if idx == 0:
            counts = {
                "Profit": len(df[df["Result"]=="Profit"]),
                "Loss":   len(df[df["Result"]=="Loss"]),
                "CTC":    len(df[df["Result"]=="CTC"]),
            }
            counts = {k:v for k,v in counts.items() if v > 0}
            if not counts:
                self._no_data_pie("No Profit/Loss/CTC data found"); return
            colors = {"Profit":self.PROFIT_C,"Loss":self.LOSS_C,"CTC":self.CTC_C}
            labels = list(counts.keys())
            values = list(counts.values())
            clrs   = [colors[l] for l in labels]
            ax = self.pie_fig.add_subplot(111)
            ax.set_facecolor(self.BG_FIG)
            wedges, texts, autotexts = ax.pie(
                values, labels=None, colors=clrs,
                autopct="%1.1f%%", startangle=140,
                pctdistance=0.78,
                wedgeprops={"linewidth":2,"edgecolor":"white","width":0.55}
            )
            for at in autotexts:
                at.set_fontsize(11); at.set_fontweight("bold"); at.set_color("white")
            ax.set_title("Overall Profit / Loss / CTC Ratio",
                         fontsize=13,fontweight="bold",color=self.TEXT_C,pad=12)
            for lbl,val,clr in zip(labels,values,clrs):
                tot = sum(values)
                self._add_legend_pill(self.pie_legend, clr, lbl,
                                      f"— {val}  ({val/tot*100:.1f}%)")
            self.pie_legend.addStretch()

        # ── Option 1: Entry with/without Rules ────────────────────────────
        elif idx == 1:
            # Two side-by-side pies: "Entry with Rules=Yes" and "Entry with Rules=No"
            rules_yes = df[df["Entry with Rules"]=="Yes"]
            rules_no  = df[df["Entry with Rules"]=="No"]

            def result_counts(sub):
                return {
                    "Profit": len(sub[sub["Result"]=="Profit"]),
                    "Loss":   len(sub[sub["Result"]=="Loss"]),
                    "CTC":    len(sub[sub["Result"]=="CTC"]),
                }

            cy = result_counts(rules_yes)
            cn = result_counts(rules_no)
            colors = {"Profit":self.PROFIT_C,"Loss":self.LOSS_C,"CTC":self.CTC_C}

            ax1 = self.pie_fig.add_subplot(121)
            ax2 = self.pie_fig.add_subplot(122)
            self.pie_fig.patch.set_facecolor(self.BG_FIG)

            for ax, counts, ttl in [(ax1,cy,"With Rules"), (ax2,cn,"Without Rules")]:
                ax.set_facecolor(self.BG_FIG)
                filtered = {k:v for k,v in counts.items() if v > 0}
                if not filtered:
                    ax.text(0.5,0.5,"No data",ha="center",va="center",
                            fontsize=11,color=self.TEXT_C,transform=ax.transAxes)
                    ax.axis("off")
                    ax.set_title(ttl,fontsize=12,fontweight="bold",color=self.TEXT_C,pad=10)
                    continue
                lbls = list(filtered.keys())
                vals = list(filtered.values())
                clrs = [colors[l] for l in lbls]
                wedges,_,autotexts = ax.pie(
                    vals, colors=clrs, autopct="%1.1f%%", startangle=140,
                    pctdistance=0.75,
                    wedgeprops={"linewidth":2,"edgecolor":"white","width":0.55}
                )
                for at in autotexts:
                    at.set_fontsize(10); at.set_fontweight("bold"); at.set_color("white")
                ax.set_title(ttl,fontsize=12,fontweight="bold",color=self.TEXT_C,pad=10)

            for lbl,clr in [("Profit",self.PROFIT_C),("Loss",self.LOSS_C),("CTC",self.CTC_C)]:
                y_v = cy.get(lbl,0); n_v = cn.get(lbl,0)
                self._add_legend_pill(self.pie_legend, clr, lbl,
                                      f"— Rules: {y_v}  /  No Rules: {n_v}")
            self.pie_legend.addStretch()

        # ── Option 2: Trading vs Non-Trading Days ─────────────────────────
        elif idx == 2:
            yes_count = len(df[df["Entry "].str.strip()=="Yes"])
            no_count  = len(df[df["Entry "].str.strip()=="No"])
            total     = yes_count + no_count
            if total == 0:
                self._no_data_pie("No Entry data found"); return

            ax = self.pie_fig.add_subplot(111)
            ax.set_facecolor(self.BG_FIG)
            values = [yes_count, no_count]
            clrs   = [self.YES_C, self.NO_C]
            labels = ["Trading Days (Yes)", "Non-Trading Days (No)"]
            wedges,_,autotexts = ax.pie(
                values, colors=clrs, autopct="%1.1f%%", startangle=90,
                pctdistance=0.78,
                wedgeprops={"linewidth":2,"edgecolor":"white","width":0.55}
            )
            for at in autotexts:
                at.set_fontsize(11); at.set_fontweight("bold"); at.set_color("white")
            ax.set_title("Trading Days vs Non-Trading Days",
                         fontsize=13,fontweight="bold",color=self.TEXT_C,pad=12)
            for lbl,val,clr in zip(labels,values,clrs):
                self._add_legend_pill(self.pie_legend, clr, lbl,
                                      f"— {val}  ({val/total*100:.1f}%)")
            self.pie_legend.addStretch()

        self.pie_canvas.draw()

    def _no_data_pie(self, msg):
        ax = self.pie_fig.add_subplot(111)
        ax.set_facecolor(self.BG_AX)
        ax.text(0.5,0.5,msg,ha="center",va="center",
                fontsize=12,color=self.TEXT_C,transform=ax.transAxes)
        ax.axis("off")
        self.pie_canvas.draw()

    # ══════════════════════════════════════════════════════════════════════════
    # BAR / STATS CHART DRAWING
    # ══════════════════════════════════════════════════════════════════════════
    def _draw_bar(self):
        self.bar_fig.clear()
        df = load_df()
        idx = self.bar_combo.currentIndex()
        self._clear_legend(self.bar_legend)

        if df.empty:
            ax = self.bar_fig.add_subplot(111)
            ax.set_facecolor(self.BG_AX)
            ax.text(0.5,0.5,"No data yet — add trades to see charts",
                    ha="center",va="center",fontsize=12,color=self.TEXT_C,
                    transform=ax.transAxes)
            ax.axis("off")
            self.bar_canvas.draw()
            return

        # ── Option 0: Points per Trade coloured by Result ─────────────────
        if idx == 0:
            # Filter rows with a numeric Points value and a result
            plot_df = df[df["Result"].isin(["Profit","Loss","CTC"])].copy()
            plot_df["pts_num"] = pd.to_numeric(
                plot_df["Points Captured / Loss"], errors="coerce")
            plot_df = plot_df.dropna(subset=["pts_num"])

            if plot_df.empty:
                ax = self.bar_fig.add_subplot(111)
                ax.set_facecolor(self.BG_AX)
                ax.text(0.5,0.5,"No numeric Points data found",
                        ha="center",va="center",fontsize=12,
                        color=self.TEXT_C,transform=ax.transAxes)
                ax.axis("off")
                self.bar_canvas.draw()
                return

            ax = self.bar_fig.add_subplot(111)
            self._style_ax(ax, "Points Captured / Loss per Trade")
            color_map = {"Profit":self.PROFIT_C,"Loss":self.LOSS_C,"CTC":self.CTC_C}
            bar_colors = [color_map.get(r, C["ink3"])
                          for r in plot_df["Result"]]
            x = range(len(plot_df))
            bars = ax.bar(x, plot_df["pts_num"].values, color=bar_colors,
                          width=0.7, zorder=3)

            # Horizontal zero line
            ax.axhline(0, color=self.TEXT_C, linewidth=0.8, zorder=4)

            # Value labels on top of bars (only if not too many)
            if len(plot_df) <= 40:
                for bar,val in zip(bars, plot_df["pts_num"].values):
                    ypos = bar.get_height() + (0.5 if val >= 0 else -2.5)
                    ax.text(bar.get_x()+bar.get_width()/2, ypos,
                            f"{val:+.0f}", ha="center", va="bottom",
                            fontsize=8, color=self.TEXT_C, fontweight="bold")

            # X-axis: short date labels
            tick_labels = [str(d)[:5] for d in plot_df["Date"].values]
            ax.set_xticks(list(x))
            ax.set_xticklabels(tick_labels, rotation=60, ha="right",
                               fontsize=8, color=self.TEXT_C)
            ax.set_ylabel("Points", fontsize=10, color=self.TEXT_C)

            for lbl,clr in [("Profit",self.PROFIT_C),
                             ("Loss",self.LOSS_C),("CTC",self.CTC_C)]:
                cnt = len(plot_df[plot_df["Result"]==lbl])
                if cnt: self._add_legend_pill(self.bar_legend, clr, lbl, f"— {cnt} trades")
            self.bar_legend.addStretch()

        # ── Option 1: Weekly Performance (Day of Week) ────────────────────
        elif idx == 1:
            plot_df = df[df["Result"].isin(["Profit","Loss"])].copy()
            plot_df["pts_num"] = pd.to_numeric(
                plot_df["Points Captured / Loss"], errors="coerce").fillna(0)

            day_profit = {d: 0.0 for d in self.DAY_ORDER}
            day_loss   = {d: 0.0 for d in self.DAY_ORDER}
            day_p_cnt  = {d: 0   for d in self.DAY_ORDER}
            day_l_cnt  = {d: 0   for d in self.DAY_ORDER}

            for _, row in plot_df.iterrows():
                day = str(row["Day"]).strip()
                if day not in day_profit: continue
                if row["Result"] == "Profit":
                    day_profit[day] += row["pts_num"]
                    day_p_cnt[day]  += 1
                else:
                    day_loss[day]   += row["pts_num"]
                    day_l_cnt[day]  += 1

            # Only show days that have data
            active_days = [d for d in self.DAY_ORDER
                           if day_p_cnt[d] > 0 or day_l_cnt[d] > 0]
            if not active_days:
                ax = self.bar_fig.add_subplot(111)
                ax.set_facecolor(self.BG_AX)
                ax.text(0.5,0.5,"No Profit/Loss data with Day info found",
                        ha="center",va="center",fontsize=12,
                        color=self.TEXT_C,transform=ax.transAxes)
                ax.axis("off")
                self.bar_canvas.draw()
                return

            import numpy as np
            x  = np.arange(len(active_days))
            w  = 0.35
            ax = self.bar_fig.add_subplot(111)
            self._style_ax(ax, "Weekly Performance — Points by Day of Week")

            p_vals = [day_profit[d] for d in active_days]
            l_vals = [day_loss[d]   for d in active_days]

            bars_p = ax.bar(x - w/2, p_vals, width=w,
                            color=self.PROFIT_C, label="Profit", zorder=3)
            bars_l = ax.bar(x + w/2, l_vals, width=w,
                            color=self.LOSS_C,   label="Loss",   zorder=3)
            ax.axhline(0, color=self.TEXT_C, linewidth=0.8, zorder=4)

            # Value labels
            for bar, val in zip(list(bars_p)+list(bars_l),
                                 p_vals + l_vals):
                if val == 0: continue
                ypos = bar.get_height() + (0.3 if val >= 0 else -3)
                ax.text(bar.get_x()+bar.get_width()/2, ypos,
                        f"{val:+.0f}", ha="center", va="bottom",
                        fontsize=9, color=self.TEXT_C, fontweight="bold")

            ax.set_xticks(x)
            ax.set_xticklabels(active_days, fontsize=10, color=self.TEXT_C)
            ax.set_ylabel("Total Points", fontsize=10, color=self.TEXT_C)

            # Summary pills
            for day in active_days:
                pc = day_p_cnt[day]; lc = day_l_cnt[day]
                pv = day_profit[day]; lv = day_loss[day]
                if pc or lc:
                    summary = f"P:{pc} ({pv:+.0f})  L:{lc} ({lv:+.0f})"
                    self._add_legend_pill(self.bar_legend,
                                          self.PROFIT_C if pv+lv>=0 else self.LOSS_C,
                                          day[:3], summary)
            self.bar_legend.addStretch()

        self.bar_canvas.draw()

    # ── Public refresh ────────────────────────────────────────────────────────
    def load(self):
        self._draw_pie()
        self._draw_bar()


# ── Main Window ────────────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Trade Diary — Backtest Manager")
        self.setMinimumSize(960,700)
        self.resize(1080,780)

        if not os.path.exists(EXCEL_FILE):
            init_excel()

        central = QWidget()
        central.setStyleSheet(f"background:{C['bg']};")
        self.setCentralWidget(central)
        ml = QVBoxLayout(central); ml.setContentsMargins(0,0,0,0); ml.setSpacing(0)

        # Header
        hdr = QWidget(); hdr.setFixedHeight(62)
        hdr.setStyleSheet(f"background:{C['ink']};")
        hl = QHBoxLayout(hdr); hl.setContentsMargins(28,0,28,0)
        tl = QLabel("Trade <span style='color:#C8512A'>Diary</span>")
        tl.setTextFormat(Qt.TextFormat.RichText)
        tl.setFont(QFont("Segoe UI",17,QFont.Weight.Bold))
        tl.setStyleSheet("color:white;background:transparent;")
        sl = QLabel("BACKTEST MANAGER")
        sl.setFont(QFont("Consolas",9))
        sl.setStyleSheet(f"color:{C['ink3']};background:transparent;")
        tc = QVBoxLayout(); tc.setSpacing(1); tc.addWidget(tl); tc.addWidget(sl)
        hl.addLayout(tc); hl.addStretch()
        ml.addWidget(hdr)

        self.tabs = QTabWidget()
        self.tabs.setDocumentMode(True)
        self.tabs.setStyleSheet(f"""
            QTabWidget::pane{{border:none;background:{C['bg']};}}
            QTabBar::tab{{background:white;color:{C['ink3']};border:none;
                border-bottom:3px solid transparent;padding:12px 22px;
                font-family:Segoe UI;font-size:13px;font-weight:600;margin-right:2px;}}
            QTabBar::tab:selected{{color:{C['accent']};border-bottom:3px solid {C['accent']};}}
            QTabBar::tab:hover{{color:{C['ink']};}}
            QTabWidget>QTabBar{{background:white;border-bottom:2px solid {C['border']};}}
        """)

        self.add_tab      = AddTradeTab()
        self.find_tab     = FindTab()
        self.results_tab  = ResultsTab()
        self.ratio_tab    = RatioTab()
        self.sheet_tab    = SheetTab()
        self.analysis_tab = AnalysisTab()

        self.tabs.addTab(self.add_tab,      "➕  Add Trade")
        self.tabs.addTab(self.find_tab,     "🔍  Find by Date")
        self.tabs.addTab(self.results_tab,  "📊  Check Results")
        self.tabs.addTab(self.ratio_tab,    "⚖  P/L Ratio")
        self.tabs.addTab(self.sheet_tab,    "📋  View Sheet")
        self.tabs.addTab(self.analysis_tab, "📈  Analysis")

        self.add_tab.trade_saved.connect(self._refresh_all)
        self.find_tab.refresh_needed.connect(self._refresh_all)
        self.results_tab.refresh_needed.connect(self._refresh_all)
        self.sheet_tab.refresh_needed.connect(self._refresh_all)
        self.tabs.currentChanged.connect(self._on_tab)
        ml.addWidget(self.tabs)

    def _refresh_all(self):
        self.sheet_tab.load()
        self.ratio_tab.load()
        self.results_tab.load_results()
        self.analysis_tab.load()

    def _on_tab(self, i):
        if i==4: self.sheet_tab.load()
        if i==3: self.ratio_tab.load()
        if i==2: self.results_tab.load_results()
        if i==5: self.analysis_tab.load()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setApplicationName("Trade Diary")
    app.setStyle("Fusion")
    palette = QPalette()
    palette.setColor(QPalette.ColorRole.Window, QColor(C["bg"]))
    palette.setColor(QPalette.ColorRole.WindowText, QColor(C["ink"]))
    palette.setColor(QPalette.ColorRole.Base, QColor(C["white"]))
    app.setPalette(palette)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())
